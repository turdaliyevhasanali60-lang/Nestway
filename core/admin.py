import csv
from datetime import timedelta
from django.contrib import admin
from django.http import HttpResponse, HttpResponseRedirect
from django.urls import reverse
from django.utils import timezone
from django.utils.html import format_html
from unfold.admin import ModelAdmin, TabularInline

from .models import (
    SiteSettings, Service, ServiceFeature, BlogPost, ContactLead,
    AboutPage, NotificationEmail, Testimonial, FAQ,
    USTeamMember, Award, PartnerReview, DriverRequirement
)

@admin.register(SiteSettings)
class SiteSettingsAdmin(ModelAdmin):
    def changelist_view(self, request, extra_context=None):
        settings = SiteSettings.load()
        url = reverse('admin:core_sitesettings_change', args=[settings.pk])
        return HttpResponseRedirect(url)

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False

class ServiceFeatureInline(TabularInline):
    model = ServiceFeature
    extra = 1
    fields = ('title', 'description', 'order')

@admin.register(Service)
class ServiceAdmin(ModelAdmin):
    list_display = ('title', 'order', 'is_active')
    list_editable = ('order',)
    inlines = [ServiceFeatureInline]

@admin.register(BlogPost)
class BlogPostAdmin(ModelAdmin):
    list_display = ('title', 'category', 'is_published', 'published_at')
    list_filter = ('category', 'is_published')
    prepopulated_fields = {'slug': ('title',)}

class ReadStatusFilter(admin.SimpleListFilter):
    title = 'Section'
    parameter_name = 'status'

    def lookups(self, request, model_admin):
        return (
            ('unread', 'Unread Contacts (Only)'),
            ('all', 'All Contacts'),
        )

    def choices(self, changelist):
        for lookup, title in self.lookup_choices:
            yield {
                'selected': self.value() == lookup or (self.value() is None and lookup == 'unread'),
                'query_string': changelist.get_query_string({self.parameter_name: lookup}, []),
                'display': title,
            }

    def queryset(self, request, queryset):
        if self.value() == 'all':
            return queryset
        if self.value() is None or self.value() == 'unread':
            return queryset.filter(is_read=False)
        return queryset

class CreatedAtFilter(admin.SimpleListFilter):
    title = 'Created At (Range)'
    parameter_name = 'created_at_range'

    def lookups(self, request, model_admin):
        return (
            ('today', 'Today'),
            ('yesterday', 'Yesterday'),
            ('this_week', 'This week'),
            ('last_week', 'Last week'),
            ('this_month', 'This month'),
            ('last_month', 'Last month'),
            ('this_year', 'This year'),
        )

    def queryset(self, request, queryset):
        now = timezone.now()
        today = now.date()
        
        if self.value() == 'today':
            return queryset.filter(created_at__date=today)
            
        elif self.value() == 'yesterday':
            yesterday = today - timedelta(days=1)
            return queryset.filter(created_at__date=yesterday)
            
        elif self.value() == 'this_week':
            start_of_week = today - timedelta(days=today.weekday())
            return queryset.filter(created_at__date__gte=start_of_week)
            
        elif self.value() == 'last_week':
            start_of_this_week = today - timedelta(days=today.weekday())
            start_of_last_week = start_of_this_week - timedelta(days=7)
            end_of_last_week = start_of_this_week - timedelta(days=1)
            return queryset.filter(created_at__date__range=[start_of_last_week, end_of_last_week])
            
        elif self.value() == 'this_month':
            start_of_month = today.replace(day=1)
            return queryset.filter(created_at__date__gte=start_of_month)
            
        elif self.value() == 'last_month':
            first_day_this_month = today.replace(day=1)
            last_day_last_month = first_day_this_month - timedelta(days=1)
            start_of_last_month = last_day_last_month.replace(day=1)
            return queryset.filter(created_at__date__range=[start_of_last_month, last_day_last_month])
            
        elif self.value() == 'this_year':
            start_of_year = today.replace(month=1, day=1)
            return queryset.filter(created_at__date__gte=start_of_year)
            
        return queryset


class YearFilter(admin.SimpleListFilter):
    title = 'Specific Year'
    parameter_name = 'created_at_year'

    def lookups(self, request, model_admin):
        years = model_admin.get_queryset(request).dates('created_at', 'year')
        return [(str(y.year), str(y.year)) for y in years]

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(created_at__year=self.value())
        return queryset


class MonthFilter(admin.SimpleListFilter):
    title = 'Specific Month'
    parameter_name = 'created_at_month'

    def lookups(self, request, model_admin):
        return (
            ('1', 'January'),
            ('2', 'February'),
            ('3', 'March'),
            ('4', 'April'),
            ('5', 'May'),
            ('6', 'June'),
            ('7', 'July'),
            ('8', 'August'),
            ('9', 'September'),
            ('10', 'October'),
            ('11', 'November'),
            ('12', 'December'),
        )

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(created_at__month=self.value())
        return queryset


class DayFilter(admin.SimpleListFilter):
    title = 'Specific Day'
    parameter_name = 'created_at_day'

    def lookups(self, request, model_admin):
        dates = []
        today = timezone.now().date()
        for i in range(30):
            d = today - timedelta(days=i)
            dates.append((d.isoformat(), d.strftime('%Y-%m-%d (%a)')))
        return dates

    def queryset(self, request, queryset):
        if self.value():
            return queryset.filter(created_at__date=self.value())
        return queryset


@admin.action(description="Export selected leads to Excel (.xlsx)")
def export_as_xlsx(modeladmin, request, queryset):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = Workbook()
    ws = wb.active
    ws.title = "Contact Leads"
    
    # Enable grid lines explicitly
    ws.views.sheetView[0].showGridLines = True
    
    # Styled headers & body
    header_fill = PatternFill(start_color="008B8B", end_color="008B8B", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    
    alignment_center = Alignment(horizontal="center", vertical="center")
    alignment_left = Alignment(horizontal="left", vertical="center")
    
    thin_border_side = Side(style='thin', color='D3D3D3')
    thin_border = Border(left=thin_border_side, right=thin_border_side, top=thin_border_side, bottom=thin_border_side)
    
    headers = ['Name', 'Email', 'Phone', 'Message', 'Created At']
    ws.append(headers)
    
    for col_num in range(1, 6):
        cell = ws.cell(row=1, column=col_num)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = alignment_center
        cell.border = thin_border
        
    for lead in queryset:
        local_time = timezone.localtime(lead.created_at).strftime('%Y-%m-%d %H:%M:%S')
        row_data = [lead.name, lead.email, lead.phone, lead.message, local_time]
        ws.append(row_data)
        
        current_row = ws.max_row
        for col_num in range(1, 6):
            cell = ws.cell(row=current_row, column=col_num)
            cell.font = data_font
            cell.border = thin_border
            if col_num in [3, 5]:
                cell.alignment = alignment_center
            else:
                cell.alignment = alignment_left
                
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            val = str(cell.value or '')
            if len(val) > max_len:
                max_len = len(val)
        ws.column_dimensions[col_letter].width = max(max_len + 3, 12)
        
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="contact_leads.xlsx"'
    wb.save(response)
    return response


@admin.action(description="Mark selected leads as Read")
def mark_as_read(modeladmin, request, queryset):
    queryset.update(is_read=True)

@admin.register(ContactLead)
class ContactLeadAdmin(ModelAdmin):
    list_display = ('colored_name', 'lead_type', 'email', 'phone', 'created_at', 'is_read')
    list_editable = ('is_read',)
    readonly_fields = ('name', 'email', 'phone', 'message', 'lead_type', 'country', 'experience_level', 'created_at')
    list_filter = ('lead_type', ReadStatusFilter, CreatedAtFilter, YearFilter, MonthFilter, DayFilter)
    actions = [export_as_xlsx, mark_as_read]
    
    @admin.display(description='Name', ordering='name')
    def colored_name(self, obj):
        if not obj.is_read:
            return format_html(
                '<span class="font-bold text-red-600 dark:text-red-400">{} '
                '<span class="ml-2 inline-flex items-center rounded bg-red-100 px-2 py-0.5 text-xs font-medium text-red-800 dark:bg-red-900 dark:text-red-200">Unread</span>'
                '</span>',
                obj.name
            )
        return obj.name
    
    def has_add_permission(self, request):
        return False
        
    def has_change_permission(self, request, obj=None):
        return True

@admin.register(NotificationEmail)
class NotificationEmailAdmin(ModelAdmin):
    list_display = ('display_name', 'email', 'is_active', 'created_at')
    list_editable = ('is_active',)
    list_filter = ('is_active',)
    search_fields = ('label', 'email')
    fields = ('label', 'email', 'is_active')

    @admin.display(description='Label / Email')
    def display_name(self, obj):
        return obj.label if obj.label else '—'

@admin.register(AboutPage)
class AboutPageAdmin(ModelAdmin):
    fieldsets = (
        ('Hero Section', {
            'fields': ('hero_title', 'hero_subtitle'),
        }),
        ('Mission, Vision & Values', {
            'fields': ('mission_title', 'mission_body', 'vision_title', 'vision_body', 'values_title', 'values_body'),
        }),
        ('CEO Section', {
            'fields': ('ceo_name', 'ceo_title', 'ceo_bio', 'ceo_image'),
        }),
        ('Team Section', {
            'fields': ('team_section_title', 'team_section_body', 'team_image'),
        }),
        ('Stats', {
            'fields': (
                'stat_1_number', 'stat_1_label',
                'stat_2_number', 'stat_2_label',
                'stat_3_number', 'stat_3_label',
                'stat_4_number', 'stat_4_label',
            ),
        }),
    )

    def changelist_view(self, request, extra_context=None):
        about = AboutPage.load()
        url = reverse('admin:core_aboutpage_change', args=[about.pk])
        return HttpResponseRedirect(url)

    def has_add_permission(self, request):
        return False

    def has_delete_permission(self, request, obj=None):
        return False


@admin.register(Testimonial)
class TestimonialAdmin(ModelAdmin):
    list_display = ('author_name', 'author_subtitle', 'initials', 'order', 'is_active')
    list_editable = ('order', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('author_name', 'quote')
    fields = ('quote', 'author_name', 'author_subtitle', 'initials', 'order', 'is_active')


@admin.register(FAQ)
class FAQAdmin(ModelAdmin):
    list_display = ('question', 'order', 'is_active')
    list_editable = ('order', 'is_active')
    list_filter = ('is_active',)
    search_fields = ('question', 'answer')
    fields = ('question', 'answer', 'order', 'is_active')


@admin.register(USTeamMember)
class USTeamMemberAdmin(ModelAdmin):
    list_display = ('name', 'role', 'order', 'is_active')
    list_editable = ('order', 'is_active')
    search_fields = ('name', 'role', 'bio')
    fields = ('name', 'role', 'bio', 'image', 'order', 'is_active')


@admin.register(Award)
class AwardAdmin(ModelAdmin):
    list_display = ('title', 'issuer', 'year', 'order', 'is_active')
    list_editable = ('order', 'is_active')
    search_fields = ('title', 'issuer', 'description')
    fields = ('title', 'issuer', 'year', 'description', 'icon_name', 'order', 'is_active')


@admin.register(PartnerReview)
class PartnerReviewAdmin(ModelAdmin):
    list_display = ('author_name', 'company_name', 'relationship', 'initials', 'order', 'is_active')
    list_editable = ('order', 'is_active')
    search_fields = ('author_name', 'company_name', 'quote')
    fields = ('quote', 'author_name', 'company_name', 'relationship', 'initials', 'order', 'is_active')


@admin.register(DriverRequirement)
class DriverRequirementAdmin(ModelAdmin):
    list_display = ('requirement_text', 'driver_type', 'order', 'is_active')
    list_editable = ('order', 'is_active')
    list_filter = ('driver_type', 'is_active')
    search_fields = ('requirement_text',)
    fields = ('requirement_text', 'driver_type', 'order', 'is_active')
